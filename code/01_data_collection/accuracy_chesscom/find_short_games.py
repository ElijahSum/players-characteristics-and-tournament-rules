#!/usr/bin/env python3
"""Filter Chess.com games with very short PGNs from a workbook of missed links."""

from __future__ import annotations

import argparse
import csv
import json
import re
import subprocess
import sys
import urllib.error
import urllib.parse
import urllib.request
from dataclasses import dataclass, field
from datetime import date, datetime, time
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook

DEFAULT_INPUT = Path("short_games/new_missed_links_2026.xlsx")
DEFAULT_FILTERED_OUTPUT = Path("short_games/new_missed_links_2026_filtered.xlsx")
DEFAULT_TOO_SHORT_OUTPUT = Path("short_games/too_short_games.csv")
DEFAULT_MAX_PLIES = 2
DEFAULT_TIMEOUT_SECONDS = 30
DEFAULT_PROGRESS_EVERY = 500
DEFAULT_ARCHIVE_PROGRESS_EVERY = 100
DEFAULT_CHECKPOINT_EVERY = 100

GAME_ID_RE = re.compile(r"/game/live/(\d+)")
CURRENT_POSITION_RE = re.compile(r'\[CurrentPosition\s+"([^"]+)"\]')
HEADER_TAG_RE = re.compile(r"\[[^\]]*\]\s*")
COMMENT_RE = re.compile(r"\{[^{}]*\}")
LINE_COMMENT_RE = re.compile(r";[^\r\n]*")
NAG_RE = re.compile(r"\$\d+")
MOVE_NUMBER_RE = re.compile(r"\d+\.(?:\.\.)?")
RESULT_TOKENS = {"1-0", "0-1", "1/2-1/2", "*"}
DATE_FORMATS = (
    "%Y-%m-%d %H:%M:%S",
    "%Y-%m-%d",
    "%b %d, %Y, %I:%M %p",
    "%b %d, %Y",
)


@dataclass(frozen=True)
class ArchiveKey:
    player: str
    year: int
    month: int


@dataclass
class ArchiveResult:
    key: ArchiveKey
    games_by_id: dict[str, dict[str, Any]]
    error: str | None = None

    @property
    def ok(self) -> bool:
        return self.error is None


@dataclass
class RunStats:
    input_rows: int = 0
    archives_fetched: int = 0
    api_failures: int = 0
    matched_rows: int = 0
    fallback_matches: int = 0
    too_short_rows: int = 0
    retained_rows: int = 0
    no_match_rows: int = 0
    unparseable_pgn_rows: int = 0
    api_error_examples: list[str] = field(default_factory=list)


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description=(
            "Scan Chess.com monthly archives for workbook rows, keep only rows with "
            "more than N plies, and save short games separately."
        )
    )
    parser.add_argument("--input", type=Path, default=DEFAULT_INPUT)
    parser.add_argument("--filtered-output", type=Path, default=DEFAULT_FILTERED_OUTPUT)
    parser.add_argument("--too-short-output", type=Path, default=DEFAULT_TOO_SHORT_OUTPUT)
    parser.add_argument("--max-plies", type=int, default=DEFAULT_MAX_PLIES)
    parser.add_argument("--timeout", type=int, default=DEFAULT_TIMEOUT_SECONDS)
    parser.add_argument("--progress-every", type=int, default=DEFAULT_PROGRESS_EVERY)
    parser.add_argument(
        "--archive-progress-every",
        type=int,
        default=DEFAULT_ARCHIVE_PROGRESS_EVERY,
    )
    parser.add_argument("--checkpoint-every", type=int, default=DEFAULT_CHECKPOINT_EVERY)
    return parser.parse_args()


def extract_game_id(url: Any) -> str | None:
    if url is None:
        return None

    match = GAME_ID_RE.search(str(url).strip())
    if not match:
        return None
    return match.group(1)


def parse_excel_datetime(value: Any) -> datetime:
    if isinstance(value, datetime):
        return value
    if isinstance(value, date):
        return datetime.combine(value, time.min)

    text = str(value).strip()
    if not text:
        raise ValueError("date is blank")

    try:
        return datetime.fromisoformat(text)
    except ValueError:
        pass

    for fmt in DATE_FORMATS:
        try:
            return datetime.strptime(text, fmt)
        except ValueError:
            continue

    raise ValueError(f"unsupported date format: {value!r}")


def normalize_player_name(value: Any) -> str:
    text = str(value).strip().lower()
    if not text:
        raise ValueError("player name is blank")
    return text


def count_plies_from_current_position(pgn: str | None) -> int | None:
    if not pgn:
        return None

    match = CURRENT_POSITION_RE.search(pgn)
    if not match:
        return None

    fen_parts = match.group(1).strip().split()
    if len(fen_parts) < 6:
        return None

    active_color = fen_parts[1]
    fullmove_text = fen_parts[5]

    try:
        fullmove_number = int(fullmove_text)
    except ValueError:
        return None

    if fullmove_number < 1 or active_color not in {"w", "b"}:
        return None

    return (fullmove_number - 1) * 2 + (1 if active_color == "b" else 0)


def count_plies_from_movetext(pgn: str | None) -> int | None:
    if not pgn:
        return None

    movetext = HEADER_TAG_RE.sub(" ", pgn)
    movetext = COMMENT_RE.sub(" ", movetext)
    movetext = LINE_COMMENT_RE.sub(" ", movetext)
    movetext = NAG_RE.sub(" ", movetext)

    previous = None
    while movetext != previous:
        previous = movetext
        movetext = re.sub(r"\([^()]*\)", " ", movetext)

    tokens: list[str] = []
    for token in movetext.split():
        if token in RESULT_TOKENS:
            continue
        if MOVE_NUMBER_RE.fullmatch(token):
            continue
        if token == "...":
            continue
        tokens.append(token)

    if not tokens:
        return 0

    return len(tokens)


def get_pgn_ply_count(pgn: str | None) -> int | None:
    ply_count = count_plies_from_current_position(pgn)
    if ply_count is not None:
        return ply_count
    return count_plies_from_movetext(pgn)


def build_archive_url(key: ArchiveKey) -> str:
    player = urllib.parse.quote(key.player, safe="")
    return f"https://api.chess.com/pub/player/{player}/games/{key.year}/{key.month:02d}"


def fetch_archive_with_urllib(url: str, timeout_seconds: int) -> dict[str, Any]:
    request = urllib.request.Request(
        url,
        headers={
            "User-Agent": "chess-accuracy-parsing/find_short_games.py",
            "Accept": "application/json",
        },
    )

    with urllib.request.urlopen(request, timeout=timeout_seconds) as response:
        charset = response.headers.get_content_charset() or "utf-8"
        payload = response.read().decode(charset)
    return json.loads(payload)


def fetch_archive_with_curl(url: str, timeout_seconds: int) -> dict[str, Any]:
    completed = subprocess.run(
        [
            "curl",
            "--silent",
            "--show-error",
            "--fail",
            "--location",
            "--compressed",
            "--connect-timeout",
            str(timeout_seconds),
            "--max-time",
            str(timeout_seconds),
            "--user-agent",
            "chess-accuracy-parsing/find_short_games.py",
            "--header",
            "Accept: application/json",
            url,
        ],
        capture_output=True,
        check=True,
        text=True,
    )
    return json.loads(completed.stdout)


def fetch_archive(
    key: ArchiveKey,
    timeout_seconds: int,
    archive_progress_every: int,
    cache: dict[ArchiveKey, ArchiveResult],
    stats: RunStats,
) -> ArchiveResult:
    cached = cache.get(key)
    if cached is not None:
        return cached

    url = build_archive_url(key)
    stats.archives_fetched += 1
    if archive_progress_every > 0 and stats.archives_fetched % archive_progress_every == 0:
        print(
            f"Fetched {stats.archives_fetched} archive requests...",
            file=sys.stderr,
            flush=True,
        )

    try:
        payload = fetch_archive_with_urllib(url, timeout_seconds)
    except Exception as urllib_exc:
        try:
            payload = fetch_archive_with_curl(url, timeout_seconds)
        except (
            subprocess.CalledProcessError,
            FileNotFoundError,
            json.JSONDecodeError,
        ) as curl_exc:
            stats.api_failures += 1
            if len(stats.api_error_examples) < 5:
                stats.api_error_examples.append(
                    f"{url} -> urllib: {urllib_exc}; curl: {curl_exc}"
                )
            result = ArchiveResult(
                key=key,
                games_by_id={},
                error=f"urllib: {urllib_exc}; curl: {curl_exc}",
            )
            cache[key] = result
            return result

    if not isinstance(payload, dict):
        stats.api_failures += 1
        if len(stats.api_error_examples) < 5:
            stats.api_error_examples.append(
                f"{url} -> unexpected payload type: {type(payload).__name__}"
            )
        result = ArchiveResult(
            key=key,
            games_by_id={},
            error=f"unexpected payload type: {type(payload).__name__}",
        )
        cache[key] = result
        return result

    games_by_id: dict[str, dict[str, Any]] = {}
    for game in payload.get("games", []):
        game_id = extract_game_id(game.get("url"))
        if game_id and game_id not in games_by_id:
            games_by_id[game_id] = game

    result = ArchiveResult(key=key, games_by_id=games_by_id)
    cache[key] = result
    return result


def write_filtered_workbook(
    sheet_name: str,
    header: list[Any],
    rows: list[list[Any]],
    path: Path,
) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)

    workbook = Workbook(write_only=True)
    worksheet = workbook.create_sheet(title=sheet_name)
    worksheet.append(header)
    for row in rows:
        worksheet.append(row)
    workbook.save(path)


def write_too_short_csv(
    header: list[Any],
    rows: list[list[Any]],
    path: Path,
) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    csv_header = get_too_short_csv_header(header)

    with path.open("w", newline="", encoding="utf-8") as handle:
        writer = csv.writer(handle)
        writer.writerow(csv_header)
        writer.writerows(rows)


def get_too_short_csv_header(header: list[Any]) -> list[str]:
    csv_header = ["" if value is None else str(value) for value in header]
    csv_header.extend(["game_id", "matched_game_url", "matched_archive_player", "ply_count"])
    return csv_header


def build_checkpoint_path(path: Path) -> Path:
    return path.with_name(f"{path.stem}.partial{path.suffix}")


def print_summary(
    stats: RunStats,
    filtered_output: Path,
    too_short_output: Path,
    checkpoint_output: Path,
) -> None:
    print("Finished filtering short games.")
    print(f"Input rows: {stats.input_rows}")
    print(f"Archives fetched: {stats.archives_fetched}")
    print(f"API failures: {stats.api_failures}")
    print(f"Matched rows: {stats.matched_rows}")
    print(f"Fallback matches: {stats.fallback_matches}")
    print(f"Too-short rows: {stats.too_short_rows}")
    print(f"Retained rows: {stats.retained_rows}")
    print(f"No-match rows kept: {stats.no_match_rows}")
    print(f"Unparseable-PGN rows kept: {stats.unparseable_pgn_rows}")
    for idx, error in enumerate(stats.api_error_examples, start=1):
        print(f"API failure sample {idx}: {error}")
    print(f"Filtered workbook: {filtered_output}")
    print(f"Too-short CSV: {too_short_output}")
    print(f"Checkpoint CSV: {checkpoint_output}")


def retain_row(filtered_rows: list[list[Any]], row_values: list[Any], stats: RunStats) -> None:
    stats.retained_rows += 1
    filtered_rows.append(row_values)


def main() -> int:
    args = parse_args()

    if args.max_plies < 0:
        print("--max-plies must be >= 0", file=sys.stderr)
        return 2

    print(f"Opening workbook: {args.input}", file=sys.stderr, flush=True)

    workbook = load_workbook(args.input, read_only=True, data_only=True)
    sheet_name = workbook.sheetnames[0]
    worksheet = workbook[sheet_name]
    row_iter = worksheet.iter_rows(values_only=True)

    try:
        raw_header = next(row_iter)
    except StopIteration:
        print(f"Input workbook is empty: {args.input}", file=sys.stderr)
        return 1

    header = list(raw_header)
    column_positions = {
        str(name).strip(): idx
        for idx, name in enumerate(header)
        if name is not None and str(name).strip()
    }

    required_columns = ("white_name", "black_name", "date", "links")
    missing = [name for name in required_columns if name not in column_positions]
    if missing:
        print(
            f"Input workbook is missing required columns: {', '.join(missing)}",
            file=sys.stderr,
        )
        return 1

    filtered_rows: list[list[Any]] = []
    too_short_rows: list[list[Any]] = []
    archive_cache: dict[ArchiveKey, ArchiveResult] = {}
    stats = RunStats()
    checkpoint_output = build_checkpoint_path(args.too_short_output)
    checkpoint_output.parent.mkdir(parents=True, exist_ok=True)

    print(
        (
            "Starting scan: rows will be checked against monthly archives, "
            f"row progress every {args.progress_every}, "
            f"archive progress every {args.archive_progress_every}, "
            f"short-game checkpoint every {args.checkpoint_every} rows."
        ),
        file=sys.stderr,
        flush=True,
    )

    white_idx = column_positions["white_name"]
    black_idx = column_positions["black_name"]
    date_idx = column_positions["date"]
    link_idx = column_positions["links"]

    with checkpoint_output.open("w", newline="", encoding="utf-8") as checkpoint_handle:
        checkpoint_writer = csv.writer(checkpoint_handle)
        checkpoint_writer.writerow(get_too_short_csv_header(header))
        checkpoint_handle.flush()

        for row_number, row in enumerate(row_iter, start=2):
            row_values = list(row)
            if not any(value is not None and str(value).strip() != "" for value in row_values):
                continue

            stats.input_rows += 1

            if args.progress_every > 0 and stats.input_rows % args.progress_every == 0:
                print(f"Processed {stats.input_rows} rows...", file=sys.stderr, flush=True)
            if args.checkpoint_every > 0 and stats.input_rows % args.checkpoint_every == 0:
                checkpoint_handle.flush()
                print(
                    f"Checkpointed short-game CSV at row {stats.input_rows}: {checkpoint_output}",
                    file=sys.stderr,
                    flush=True,
                )

            row_should_keep = True

            try:
                white_player = normalize_player_name(row_values[white_idx])
                black_player = normalize_player_name(row_values[black_idx])
                dt = parse_excel_datetime(row_values[date_idx])
                game_id = extract_game_id(row_values[link_idx])
                if game_id is None:
                    raise ValueError("could not extract game id from links")
            except ValueError:
                stats.no_match_rows += 1
                retain_row(filtered_rows, row_values, stats)
                continue

            matched_game: dict[str, Any] | None = None
            matched_archive_player: str | None = None

            for player in (white_player, black_player):
                archive = fetch_archive(
                    ArchiveKey(player=player, year=dt.year, month=dt.month),
                    timeout_seconds=args.timeout,
                    archive_progress_every=args.archive_progress_every,
                    cache=archive_cache,
                    stats=stats,
                )
                if not archive.ok:
                    continue

                candidate = archive.games_by_id.get(game_id)
                if candidate is None:
                    continue

                matched_game = candidate
                matched_archive_player = player
                if player == black_player:
                    stats.fallback_matches += 1
                break

            if matched_game is None or matched_archive_player is None:
                stats.no_match_rows += 1
                retain_row(filtered_rows, row_values, stats)
                continue

            stats.matched_rows += 1

            ply_count = get_pgn_ply_count(matched_game.get("pgn"))
            if ply_count is None:
                stats.unparseable_pgn_rows += 1
                retain_row(filtered_rows, row_values, stats)
                continue

            if ply_count <= args.max_plies:
                row_should_keep = False
                stats.too_short_rows += 1
                too_short_row = (
                    row_values
                    + [
                        game_id,
                        matched_game.get("url", ""),
                        matched_archive_player,
                        ply_count,
                    ]
                )
                too_short_rows.append(too_short_row)
                checkpoint_writer.writerow(too_short_row)

            if row_should_keep:
                retain_row(filtered_rows, row_values, stats)

        checkpoint_handle.flush()

    write_filtered_workbook(sheet_name, header, filtered_rows, args.filtered_output)
    write_too_short_csv(header, too_short_rows, args.too_short_output)
    print_summary(stats, args.filtered_output, args.too_short_output, checkpoint_output)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
